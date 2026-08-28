#!/usr/bin/env python3
"""Compare two HR pipeline XLSX files by workbook semantics, not ZIP packaging.

The HR pipeline writes XLSX files through openpyxl. Re-saving a workbook can
legitimately rewrite XML ordering, relationship IDs, style-table internals and
ZIP metadata even when the workbook a user sees is unchanged. This gate keeps
CI strict by comparing the workbook contract that matters: sheet order,
non-empty cell values/formulas, number formats, hyperlinks, merged ranges,
freeze panes, autofilters and sheet visibility.

If a mismatch exists, the script prints concrete sheet/cell differences before
returning a non-zero exit status and preserves the rebuilt workbook in the
existing CI diagnostics artifact for review/re-baselining.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import math
import shutil
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

MAX_DIFFS = 40


def canonical_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return {"type": "datetime", "value": value.isoformat(timespec="microseconds")}
    if isinstance(value, dt.date):
        return {"type": "date", "value": value.isoformat()}
    if isinstance(value, dt.time):
        return {"type": "time", "value": value.isoformat(timespec="microseconds")}
    if isinstance(value, float):
        if math.isnan(value):
            return {"type": "float", "value": "NaN"}
        if math.isinf(value):
            return {"type": "float", "value": "Infinity" if value > 0 else "-Infinity"}
        # 17 significant digits round-trips an IEEE-754 double without hiding
        # meaningful numeric drift.
        return {"type": "float", "value": format(value, ".17g")}
    if isinstance(value, (int, bool, str)):
        return value
    return {"type": type(value).__name__, "value": str(value)}


def freeze_value(value: Any) -> str | None:
    if value is None:
        return None
    return getattr(value, "coordinate", None) or str(value)


def cell_payload(cell: Any) -> dict[str, Any] | None:
    if cell.value is None and cell.hyperlink is None:
        return None
    hyperlink = None
    if cell.hyperlink is not None:
        hyperlink = cell.hyperlink.target or cell.hyperlink.location or str(cell.hyperlink)
    return {
        "value": canonical_value(cell.value),
        "data_type": cell.data_type,
        "number_format": cell.number_format,
        "hyperlink": hyperlink,
    }


def non_empty_cells(ws: Any) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows():
        for cell in row:
            payload = cell_payload(cell)
            if payload is not None:
                out[cell.coordinate] = payload
    return out


def sheet_metadata(ws: Any) -> dict[str, Any]:
    return {
        "sheet_state": ws.sheet_state,
        "freeze_panes": freeze_value(ws.freeze_panes),
        "auto_filter": ws.auto_filter.ref,
        "merged_ranges": sorted(str(item) for item in ws.merged_cells.ranges),
    }


def semantic_digest(workbook: Any) -> str:
    digest = hashlib.sha256()
    for ws in workbook.worksheets:
        metadata = sheet_metadata(ws)
        digest.update(json.dumps(["sheet", ws.title, metadata], ensure_ascii=False, sort_keys=True).encode("utf-8"))
        digest.update(b"\n")
        for coordinate, payload in sorted(non_empty_cells(ws).items()):
            digest.update(
                json.dumps([coordinate, payload], ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
            )
            digest.update(b"\n")
    return digest.hexdigest()


def compare_workbooks(before_path: Path, after_path: Path) -> list[str]:
    before = load_workbook(before_path, data_only=False, read_only=False)
    after = load_workbook(after_path, data_only=False, read_only=False)
    diffs: list[str] = []

    try:
        if before.sheetnames != after.sheetnames:
            diffs.append(f"sheet order/name drift: {before.sheetnames!r} != {after.sheetnames!r}")

        common = [name for name in before.sheetnames if name in after.sheetnames]
        for name in common:
            if len(diffs) >= MAX_DIFFS:
                break
            left = before[name]
            right = after[name]

            left_meta = sheet_metadata(left)
            right_meta = sheet_metadata(right)
            if left_meta != right_meta:
                diffs.append(f"{name}: sheet metadata drift: {left_meta!r} != {right_meta!r}")
                if len(diffs) >= MAX_DIFFS:
                    break

            left_cells = non_empty_cells(left)
            right_cells = non_empty_cells(right)
            for coordinate in sorted(set(left_cells) | set(right_cells)):
                if left_cells.get(coordinate) != right_cells.get(coordinate):
                    diffs.append(
                        f"{name}!{coordinate}: {left_cells.get(coordinate)!r} != {right_cells.get(coordinate)!r}"
                    )
                    if len(diffs) >= MAX_DIFFS:
                        break

        before_digest = semantic_digest(before)
        after_digest = semantic_digest(after)
        print(f"baseline semantic SHA-256: {before_digest}")
        print(f"rebuilt  semantic SHA-256: {after_digest}")
        if not diffs and before_digest != after_digest:
            diffs.append("semantic digest differs despite no enumerated sheet/cell mismatch")
        return diffs
    finally:
        before.close()
        after.close()


def preserve_rebuilt_workbook(rebuilt: Path) -> Path:
    diagnostics = Path("artifacts/diagnostics")
    diagnostics.mkdir(parents=True, exist_ok=True)
    destination = diagnostics / "hr-rebuilt-production.xlsx"
    shutil.copy2(rebuilt, destination)
    return destination


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("baseline", type=Path)
    parser.add_argument("rebuilt", type=Path)
    args = parser.parse_args()

    if not args.baseline.is_file():
        print(f"baseline workbook missing: {args.baseline}", file=sys.stderr)
        return 2
    if not args.rebuilt.is_file():
        print(f"rebuilt workbook missing: {args.rebuilt}", file=sys.stderr)
        return 2

    diffs = compare_workbooks(args.baseline, args.rebuilt)
    if diffs:
        preserved = preserve_rebuilt_workbook(args.rebuilt)
        print(f"rebuilt workbook preserved for CI review: {preserved}", file=sys.stderr)
        print(f"HR workbook semantic drift detected ({len(diffs)} reported difference(s)): ", file=sys.stderr)
        for diff in diffs:
            print(f"  - {diff}", file=sys.stderr)
        if len(diffs) >= MAX_DIFFS:
            print(f"  - output capped at {MAX_DIFFS} differences", file=sys.stderr)
        return 1

    print("HR workbook semantics match the committed baseline.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
